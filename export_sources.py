"""Export the Sources worksheet from data/sources.xlsx to sources.csv."""
from __future__ import annotations

import argparse
import csv
import os
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = BASE_DIR / "data" / "sources.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "data" / "sources.csv"
SHEET_NAME = "Sources"
SOURCE_COLUMNS = (
    "source_id",
    "source",
    "sport",
    "competition",
    "enabled",
    "source_type",
)
VALID_SOURCE_TYPES = {"ics", "wpbl_api", "wsl_official"}


def cell_text(value: Any) -> str:
    """Convert a spreadsheet value to the text expected by sources.csv."""
    if value is None:
        return ""

    if isinstance(value, bool):
        return "true" if value else "false"

    return str(value).strip()


def read_source_rows(workbook_path: Path) -> list[dict[str, str]]:
    """Read and validate source rows from an Excel workbook."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        matching_sheets = []

        for candidate in workbook.worksheets:
            first_row = next(candidate.iter_rows(values_only=True), ())
            candidate_headers = tuple(
                cell_text(value).lower() for value in first_row
            )

            if candidate_headers == SOURCE_COLUMNS:
                matching_sheets.append(candidate)

        if len(matching_sheets) == 1:
            sheet = matching_sheets[0]
            print(
                f"Using worksheet {sheet.title!r}; rename it to "
                f"{SHEET_NAME!r} for clarity."
            )
        else:
            matching_names = ", ".join(
                sheet.title for sheet in matching_sheets
            ) or "none"
            raise ValueError(
                f"{workbook_path} has no {SHEET_NAME!r} worksheet and "
                "could not identify exactly one compatible worksheet "
                f"(matches: {matching_names})."
            )

    if sheet.max_row < 1:
        raise ValueError(
            f"The {sheet.title!r} worksheet is empty."
        )

    values = sheet.iter_rows(values_only=True)
    header_values = next(values, None)

    if header_values is None:
        raise ValueError(f"The {sheet.title!r} worksheet is empty.")

    headers = tuple(cell_text(value).lower() for value in header_values)

    if headers != SOURCE_COLUMNS:
        expected = ", ".join(SOURCE_COLUMNS)
        actual = ", ".join(headers)
        raise ValueError(
            f"Unexpected columns in {sheet.title!r}.\n"
            f"Expected: {expected}\nActual: {actual}"
        )

    source_rows: list[dict[str, str]] = []
    seen_source_ids: set[str] = set()

    for row_number, values_row in enumerate(values, start=2):
        row = {
            column: cell_text(value)
            for column, value in zip(SOURCE_COLUMNS, values_row)
        }

        if not any(row.values()):
            continue

        missing = [
            column
            for column in SOURCE_COLUMNS[:4]
            if not row[column]
        ]

        if missing:
            raise ValueError(
                f"Row {row_number} is missing: {', '.join(missing)}."
            )

        source_id = row["source_id"]

        if source_id in seen_source_ids:
            raise ValueError(
                f"Duplicate source_id {source_id!r} on row {row_number}."
            )

        seen_source_ids.add(source_id)
        enabled = row["enabled"].lower() or "true"

        if enabled not in {"true", "false"}:
            raise ValueError(
                f"Row {row_number} has invalid enabled value "
                f"{row['enabled']!r}; use true or false."
            )

        source_type = row["source_type"].lower() or "ics"

        if source_type not in VALID_SOURCE_TYPES:
            choices = ", ".join(sorted(VALID_SOURCE_TYPES))
            raise ValueError(
                f"Row {row_number} has invalid source_type "
                f"{source_type!r}; use one of: {choices}."
            )

        row["enabled"] = enabled
        row["source_type"] = source_type
        source_rows.append(row)

    if not source_rows:
        raise ValueError(f"The {sheet.title!r} worksheet has no sources.")

    return source_rows


def write_sources_csv(rows: list[dict[str, str]], output_path: Path) -> None:
    """Write validated rows atomically so a failed export keeps the old CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=output_path.parent,
        prefix=f".{output_path.name}.",
        suffix=".tmp",
        text=True,
    )

    try:
        with os.fdopen(descriptor, "w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(
                file,
                fieldnames=SOURCE_COLUMNS,
                lineterminator="\n",
            )
            writer.writeheader()
            writer.writerows(rows)

        Path(temporary_name).replace(output_path)
    except BaseException:
        Path(temporary_name).unlink(missing_ok=True)
        raise


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export the Her Match sources workbook to CSV."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook path (default: {DEFAULT_WORKBOOK})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"CSV output path (default: {DEFAULT_OUTPUT})",
    )
    arguments = parser.parse_args()
    rows = read_source_rows(arguments.workbook)
    write_sources_csv(rows, arguments.output)
    print(f"Exported {len(rows)} source(s) to {arguments.output}.")


if __name__ == "__main__":
    main()
